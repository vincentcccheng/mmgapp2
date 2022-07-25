# -*- coding: utf-8 -*-

import uuid
import requests
from flask import Flask, render_template, session, request, redirect, url_for, jsonify, send_from_directory, send_file
from flask_session import Session  # https://pythonhosted.org/Flask-Session
import msal
import json
from jinja2 import Template
from flask_cors import CORS, cross_origin
from datetime import datetime, timedelta
import os 
from dotenv import load_dotenv
load_dotenv()

#below are additionals for React JS
import pymongo
from pymongo import MongoClient
from pymongo.errors import OperationFailure

#import urllib   # may be needed for filling space in URL  
from werkzeug.utils import secure_filename
import gridfs
import io
from bson.objectid import ObjectId ## required for creating an object 
from datetime import date

#newly added  # pip for Excel  
#from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from openpyxl import load_workbook
from io import BytesIO
from openpyxl.styles import Alignment
from openpyxl.workbook.protection import WorkbookProtection

# added for local time 
from datetime import datetime, timezone

# for Excel Reporting
import itertools
from itertools import groupby
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
from openpyxl.cell import MergedCell


# Below for SahrePoint
from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.files.file import File
from office365.sharepoint.listitems.caml.caml_query import CamlQuery  
from office365.runtime.http.request_options import RequestOptions
from office365.sharepoint.files.file_creation_information import FileCreationInformation


#calcuate size
#pip install python-dateutil
import math 

# parse date 
from dateutil import parser

#map maker
import folium
from folium.features import DivIcon

# for image and PDF processing
from borb.pdf.document import Document
from borb.pdf.page.page import Page
from borb.pdf.canvas.layout.image.image import Image as BorbImage 
from borb.pdf.canvas.layout.page_layout.multi_column_layout import SingleColumnLayout
from borb.pdf.canvas.layout.text.paragraph import Paragraph
from borb.pdf.canvas.layout.layout_element import Alignment  
from decimal import Decimal
from borb.pdf.pdf import PDF
from io import BytesIO
from PIL import Image, ImageFont, ImageDraw

# for zipping files
from zipfile import ZipFile

# for testing timer 
import time 

# for past inspection result  
import dateutil.parser
import datetime as datetime2

app = Flask(__name__, static_folder="frontend/build/static", template_folder="frontend/build")    #production 
    
app.config['SECRET_KEY']= os.environ['SECRET_KEY']

cluster = MongoClient(os.environ['MONGODB_URL'], tls=True, tlsAllowInvalidCertificates=True,  maxPoolSize=100)

database = os.environ['DATABASE']

db = cluster[database]

site_url = 'https://macysinc.sharepoint.com/sites/OSO/'
app_principal = {
     'client_id': os.environ['SHAREPOINT_CLIENT_ID'],
     'client_secret': os.environ['SHAREPOINT_CLIENT_SECRET'],
}

context_auth = AuthenticationContext(url=site_url)
context_auth.acquire_token_for_app(client_id=app_principal['client_id'], client_secret=app_principal['client_secret'])
    
ctx = ClientContext(site_url, context_auth)

app.config['SESSION_TYPE'] = 'mongodb'  
app.config['SESSION_KEY_PREFIX'] = 'session:' 

app.config['SESSION_MONGODB'] = cluster 
app.config['SESSION_MONGODB_DB'] =  'qcDB' 
app.config['SESSION_MONGODB_COLLECT'] = 'sessions'
#app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes = 45)
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes = int(os.environ['SESSION_TIMEOUT']))

Session(app)

app.config['UPLOAD_FOLDER'] = '.'
#app.config['MAX_CONTENT_LENGTH'] = 100 * 1024    # 100K       
app.config['MAX_CONTENT_LENGTH'] = int(os.environ['UPLOAD_MAX_SIZE'])   

# This section is needed for url_for("foo", _external=True) to automatically
# generate http scheme when this sample is running on localhost,
# and to generate https scheme when it is deployed behind reversed proxy.
# See also https://flask.palletsprojects.com/en/1.0.x/deploying/wsgi-standalone/#proxy-setups
from werkzeug.middleware.proxy_fix import ProxyFix
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)


# decorator to check whether the session is still valid 
def check_logged(f):
    def wrapper(*args, **kwargs):
        #print("enter wrapper")
        if (os.environ['ENVIRONMENT']=="PROD"):
            #print("running check_logged to check whether it is logged")
            if not (request.headers.get('api')) and not session.get("user"):
                return redirect(url_for("login"))
            if (request.headers.get('api')) and not session.get("user"):   
                return jsonify( 
                token=f'{"Invalid Sesssion"}' 
                ), 501
            print(f"current user : { session['email'].lower() }")        
        return f(*args, **kwargs)
    wrapper.__name__ = f.__name__
    return wrapper


@app.route("/api/getPhoto")
@check_logged
def getPhoto():

    #print('session email', session['email'])

    if (os.environ['ENVIRONMENT']=="PROD"):        
        print("Getting photo for production")   
        #check whether it's Macys's email account
        if  "@macys.com" not in session['email'].lower():
             return send_from_directory("frontend/public/static/img", "anonymous.jpg")        
    else:
        print("Getting photo for development")       
        return send_from_directory("frontend/public/static/img", "anonymous.jpg")
    
    token = _get_token_from_cache(json.loads(os.environ['SCOPE']))
    if not token and not os.environ:
        return redirect(url_for("login"))
    endpoint = "https://graph.microsoft.com/v1.0/me/photos/120x120/$value"

    photo_response = requests.get(  # Use token to call downstream service
        endpoint,
        headers={'Authorization': 'Bearer ' + token['access_token']},
        stream=True) 
    photo_status_code = photo_response.status_code
    if photo_status_code == 200:
       photo = photo_response.raw.read()
       return photo 
    else:        
       return  send_from_directory("frontend/build/static/img", "anonymous.jpg")
       

@app.route("/")
@check_logged
def index():    
    if not session.get("user"):
        return redirect(url_for("login"))
    return render_template('index.html', user=session["user"], version=msal.__version__)

@app.route("/login", defaults={'timeout':None}) 
@app.route("/login/<timeout>") 
def login(timeout):
    if (timeout):
        print ("Entering login process with "  + timeout)
    # Technically we could use empty list [] as scopes to do just sign in,
    # here we choose to also collect end user consent upfront
 
    session["flow"] = _build_auth_code_flow(scopes=json.loads(os.environ['SCOPE']))
    #print (session["flow"]["auth_uri"])
    #  auth_uri an be added with prompt=login to force sign in     
    return render_template("login.html", auth_url=session["flow"]["auth_uri"], version=msal.__version__, timeout_message=timeout)

@app.route(os.environ['REDIRECT_PATH'])  # Its absolute URL must match your app's redirect_uri set in AAD
def authorized():
    try:
        print("Entering " + os.environ['REDIRECT_PATH'])
        cache = _load_cache()
        result = _build_msal_app(cache=cache).acquire_token_by_auth_code_flow(
            session.get("flow", {}), request.args)
        #print("passing " + app_config.REDIRECT_PATH)
        if "error" in result:
            return render_template("auth_error.html", result=result)
        session["user"] = result.get("id_token_claims")
        # Vincent added below:
        #print ("email", json.dumps(result.get("id_token_claims")))
        #print ("email", result.get("id_token_claims").get('email'))
        session["email"] = (result.get("id_token_claims").get('email')).lower()      
        _save_cache(cache)
    except ValueError:  # Usually caused by CSRF
        pass  # Simply ignore them
        return render_template("auth_error.html", result={"error" : "Value Error", "error_description":"Not signed in yet !!"})
    return redirect(url_for("index"))

@app.route("/logout")
def logout():
    session.clear()  # Wipe out user and its token cache from session
    return redirect(  # Also logout from your tenant's web session
        os.environ['AUTHORITY'] + "/oauth2/v2.0/logout" +
        "?post_logout_redirect_uri=" + url_for("index", _external=True))


@app.route("/graphcall")
@check_logged
def graphcall():
    token = _get_token_from_cache(json.loads(os.environ['SCOPE']))
    if not token:
        return redirect(url_for("login"))
    graph_data = requests.get(  # Use token to call downstream service
        os.environ['ENDPOINT'],
        headers={'Authorization': 'Bearer ' + token['access_token']},
        ).json()
    return render_template('display.html', result=graph_data)

def _load_cache():
    cache = msal.SerializableTokenCache()
    if session.get("token_cache"):
        cache.deserialize(session["token_cache"])
    return cache

def _save_cache(cache):
    if cache.has_state_changed:
        session["token_cache"] = cache.serialize()

def _build_msal_app(cache=None, authority=None):
    return msal.ConfidentialClientApplication(
        os.environ['CLIENT_ID'], authority=authority or os.environ['AUTHORITY'],
        client_credential=os.environ['CLIENT_SECRET'], token_cache=cache)

def _build_auth_code_flow(authority=None, scopes=None):
    return _build_msal_app(authority=authority).initiate_auth_code_flow(
        scopes or [],
        redirect_uri=url_for("authorized", _external=True))

def _get_token_from_cache(scope=None):
    cache = _load_cache()  # This web app maintains one cache per session
    cca = _build_msal_app(cache=cache)
    accounts = cca.get_accounts()
    if accounts:  # So all account(s) belong to the current signed-in user
        result = cca.acquire_token_silent(scope, account=accounts[0])
        _save_cache(cache)
        return result

app.jinja_env.globals.update(_build_auth_code_flow=_build_auth_code_flow)  # Used in template

## below for Reacj JS

def getTodayDate():
    return date.today().strftime("%m/%d/%y")  ## get today's date 


ALLOWED_EXTENSIONS = set(['pdf', 'png', 'jpg', 'jpeg', 'xlsx', 'doc', 'docx', 'ppt', 'pptx', 'zip'])

def allowed_file(filename):
	return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

ALLOWED_PHOTOS = set(['png', 'jpg', 'jpeg'])
IMAGE_FORMAT = ['PNG', 'JPEG', 'JPEG']

def allowed_photos(filename):
	return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_PHOTOS

def image_format(filename):    
    ## find corresponding FORMAT that PIL IMAGE defines
    ext = filename.rsplit('.', 1)[1].upper()       
    imageFormat = ''
    count = 0                    
    for x in ALLOWED_PHOTOS:	
        if x.upper()==ext:                            
            imageFormat = IMAGE_FORMAT[count]                                           
            break
    count +=1     
    return imageFormat     
    
@app.route('/api/upload', methods=['POST'])
@check_logged
def upload_file():
        
  if request.method == 'POST':
              
            try:    
                # inspection_id = '225233-1-F' 
                # relative_url =  "2022inspRpt/SU22975MF36843"
              

                inspection_id = request.headers['inspection_id']
                relative_url =  request.headers['relative_url']
                inspection_date =  request.headers['inspection_date']
    
                #print('Upload/Inspection_id is {0}, {1}, {2}'.format(inspection_id, su_no, mf_no))                   
                # check if the post request has the files part
                if 'files[]' not in request.files:
                        #flash('No file part')
                        return "no files", 406
              
                files = request.files.getlist('files[]')      
                print('files', files)
              
                newfiles = []

                for file in files:                           
                    
                    # file.seek(0, os.SEEK_END)
                    # file_length = file.tell()
                    # print('file.length', file_length)

                    if file and allowed_file(file.filename):                                

                        mimetype = file.content_type
                        filename = file.filename    
                        
                        # 6/23/22 remove below as secure_filename removes non-ascii characters.
                        # filename = secure_filename(filename)                          

                        target_folder = ctx.web.get_folder_by_server_relative_path(relative_url)                        
                        ctx.execute_query()
                        
                        info = FileCreationInformation()                        
                        info.content = file.read()        
                        #enable below for control of each size of the file. 

                        # if  len(info.content) >  1024 * 1024 * 1:    
                        #     return f"The size of {file.filename} exceeds 2 MB limit !", 555                                                                            
                        if filename == "image.jpg":
                            filename = "img-" +  str(uuid.uuid4()) + ".jpg"                          

                        info.url = filename                                  
                        info.overwrite = True                                
                        upload_file = target_folder.files.add(info)             
                   
                        # from office import background_task                         
                        # job = q.enqueue(background_task, ctx, upload_file,inspection_id, inspection_date,session["userName"])                                                  
                        # q_len = len(q)               
                        # print(f"job : {job.result} and len : {q_len} and job is finished : {job.is_finished}")   
                        ctx.execute_query()
 
                        list_item = upload_file.listItemAllFields # get associated list item                                 
                        list_item.set_property("Inspection_x0020_ID", inspection_id)
                        list_item.set_property("Inspection_x0020_Date", inspection_date)                              
                        list_item.set_property("Last_x0020_Editor", session["userName"])
                        list_item.update()                                
                        ctx.execute_query()         
                        
                        #ctx.execute_query()              
                        #Once file is uploaded, it's metadata could be set like this                            
                        #newfiles.append({ "_id" :str(id), "enable": True, "file_name":filename, "mime_type": mimetype })                       
                        # return jsonify(newfiles),200
                    else:
                        return f"Your file type is not allowed", 500

                return "OK",200
                                
            except Exception as e:    
               print(e)
               return f"The total size of all files may exceed a size limit of {os.environ['UPLOAD_MAX_SIZE']} bytes!", 413                        

            finally:
               pass

  else:
    return "Upload Error", 406



@app.route('/download/<string:_id>', methods=['GET'])
@check_logged
def download_file(_id):
        fs = gridfs.GridFS(db)
        # _id of file example = "60b609a291e07bc60a594482"
        # below download_name and mimetype are dummy to fullfill the send_file requirements
        # changed from download_name to attachment_filename         
        return send_file(io.BytesIO(fs.get(ObjectId(_id)).read()), attachment_filename='bug2.jpg', mimetype='image/jpg')                                                                                  
        

@app.route('/downloadSharePointFile', methods=['GET'])
@check_logged
def download_sharepoint_file():
    try:
      
      file_url = request.headers['file-url'] 
      file_url = file_url.replace("'","%27%27") ## replace character ' 
      file_url = file_url.replace("#","%23")    ## replace character #  
      
      _response = File.open_binary(ctx, file_url)                 
      data = BytesIO(_response.content)               
      return send_file(data, attachment_filename='whatever.jpg', mimetype='image/jpg')                                                                 
      #return send_file(data, attachment_filename='whatever.jpg', mimetype='image/jpg')                                                                 

    except:
      return ('Error')

@app.route('/showPhoto', methods=['GET'])
@check_logged
def showPhoto():
    try:                   
         
      folder = request.args.get('folder')
      filename  = request.args.get('filename')
      file_url = "/sites/OSO/" + os.environ["SHAREPOINT_PATH"] + "/" + folder + "/" + filename
      #print("file_url" ,file_url)
      _response = File.open_binary(ctx, file_url)                 
      data = io.BytesIO(_response.content)               
      return send_file(data,  attachment_filename='bug2.jpg', mimetype='image/tiff')                                                                 
  
    except Exception as e:
      print("Error found in showPhoto",e)
      return ('Error')


@app.route('/api/getDefectTable')
@check_logged
def getDefectTable():

    try:
        print("Calling Defect Table")    
        col = db["defectTable"]
        query = {  "id" : { "$gt" :0 }}       
        defect_array = []     
        results = col.find(query)
        for result in results:
            defect_array.append(result) 
 
        return  jsonify(defect_array)
    
    except OperationFailure:
        print("Mongo Access Error")
        return 'Error'    


@app.route('/api/getCriticalDefects')
@check_logged
def getCriticalDefects():

    try:
        print("Getting Critical from Mongo")    
        col = db["criticalDefects"]
        query = {}       
        array = []     
        results = col.find(query)    
        for result in results:
            result.pop("_id")            
            array.append(result)
        return  jsonify(array)
 
    except OperationFailure:
        print("Mongo Access Error")
        return 'Error'    


@app.route('/api/checkDuplicateID', methods=['POST'])
@check_logged
def check_duplicate_inpsection_id():
    content = request.get_json() #python data 
    _id = content['_id']
    col = db["inspectionResult"]
    query =  { "_id": _id}
    exists = col.find_one(query)
    if (exists):
        return "t",200
    else:        
        return "f",201

@app.route('/api/save', methods=['POST'])
@check_logged
def save_inspection():
    content = request.get_json() #python data 

    #content = request.data # json data 
    _id = content['_id']
    checkList = content['checkList']
    items = content['items']
    itemsTotal = content['itemsTotal'] 
    poList = content['poList'] 

    defects = content['defects']
    uploads = content['uploads']
    main = content['main']
    misc = content['misc']
    update_history = content['updateHistory']    

    # get the current time and convert to string
    updated_time = datetime.now(timezone.utc).ctime()

    if (session.get("email")):
         updated_by = session.get("email")
    else:
         updated_by = "vincent.cheng@macys.com"   
    
    # Convert dictionary object into string, this is for change tracking only
    misc_str = json.dumps(misc)   
    
    #main['inspection_date'] = parser.parse(main['inspection_date'])
    main['inspection_date'] = main['inspection_date']
        
    if (update_history ==[]) :        
        update_current = { "id" :  str(uuid.uuid4()), "misc" :misc_str,  "updated_by" : updated_by, "updated_time" : updated_time, "updated_mode" : "create"}
    else:
        update_current = { "id" :  str(uuid.uuid4()), "misc" :misc_str,  "updated_by" : updated_by, "updated_time" : updated_time, "updated_mode" : "update"}

    update_history.append(update_current)    
       
    new_content = { "_id" : _id, "main" : main,  "misc" : misc,  "checkList" : checkList, "items" : items, "itemsTotal" : itemsTotal, 
    "poList" : poList,  "defects": defects,    "uploads":uploads, "update_history" : update_history }
  
    col = db["inspectionResult"]

    query =  { "_id": _id}
    exists = col.find_one(query)

    ## convert datetime string of change tracking into datetime object in Mongodb 
    for hist in update_history:
        hist['updated_time'] = parser.parse( hist['updated_time'])


    ## convert all major and minor of defect string to integer in Mongodb 
    for defect in defects:
        defect['major'] = int(defect['major'])
        defect['minor'] = int(defect['minor'])

    if (exists):
        change =  { "$set":  {  "main" : main, "misc": misc, "checkList" : checkList, "items" : items, 
        "itemsTotal" : itemsTotal, "poList" : poList, "defects":defects, "uploads" : uploads, "update_history": update_history} }    # change     
        col.update_one(query, change)
        return "ok",200
    else:        
        x = col.insert_one(new_content)
        #print(x.inserted_id)
        return "ok",200
        

def isLegitAPI(mc_no): 
    isFound = False        
    mcTable = session.get("mcTable")
    for rec in mcTable:
        if rec['mc_no'] == mc_no:
            #print ("su, mf = {0}, {1}".format(rec["su_no"], rec["mf_no"]))
            isFound = True     
            break 
            
    return (isFound)

@app.route('/api/search',methods=['POST'])
@check_logged
def search_inspection():
        
    content = request.get_json() #python data     
    ##print("Search content",content)
    _id = content['_id']

    ## removed by Vincent on 4/2/22 for the sake of accepting MC's not in MC Table
    # if not (isLegitAPI(_id['mc'])):
    #     print('Error', 406)
    #     return "Error", 406 

    col = db["inspectionResult"]

    query =  { "_id": _id}
    results = col.find_one(query)        

    if (results):
       return  jsonify(results), 200 
    else:
       return "Error", 404 


@app.route('/api/delete',methods=['POST'])
@check_logged
def delete_inspection():
        
    content = request.get_json() #python data     
    ##print("Search content",content)
     
    try:  
        
        col = db["inspectionResult"]
        delete_log_col = db["Delete_Log"]
        _id = content['_id']

        query =  { "_id": _id}

        ## add to delete_log in Mongo
        updated_by = "development@heroku" if session.get("email") == None else session.get("email")               
        existing_inspectionResult = col.find_one(query)
        delete_log_col.insert_one( { "_id": str(uuid.uuid4()), 
        "updated_by" : updated_by, "time":datetime.now(timezone.utc),"doc_type": "inspectionResult", "rec" : existing_inspectionResult })

        result = col.delete_one(query)
        if result.deleted_count > 0:
            return  "OK", 200 
        else:
            return  "Not OK", 400 
    
    #except mongoengine.errors.OperationError:           
    except OperationFailure:
        print("error")
        return "Error", 400 


@app.route('/api/searchInspByMC',methods=['POST'])
@check_logged
def searchInspMC():
        
    content = request.get_json() #python data     
    mc = content['mc'] 
 
    col = db["inspectionResult"]
    
    ## Reserverd for later use
    # search = []
    # for _filter in session['mfList']:
    #     search.append(   {  '$and': [ { 'main.su_no': { '$eq': _filter['SU'] } }, { 'main.mf_no' : { '$eq': _filter['MF'] } } ]  } )        

    query =   {
       "_id.mc" : {
       "$regex": mc,
       "$options" :'i' # case-insensitive
       } }
     
    #print(search, '$$search')     

    ##results = col.find(query).limit(5)    
    ## returns 10 at a time
    ##  to access print(rec["_id"]["mc"])    
    results = col.find(query).limit(60)     


    id_array = []
    for result in results:
        ##'id' : {uuid.uuid4()
        rec = { '_id' : result['_id'], 'main' : result['main'] }
        id_array.append(rec)        
        
    if (results):
       return  jsonify(id_array), 200 
    else:
       return "Error", 404 


@app.route('/api/getUserProfile',methods=['POST'])
@check_logged
def getUserProfile():            
    sessionData = establishSessionData()
    if (sessionData):
       return  jsonify(sessionData), 200 
    else:
       return "Error", 404 


@app.route('/api/getMCtable',methods=['POST'])
@check_logged
def getMCtable():            
  
    content = request.get_json() #python data     
    su = content['su']
    mf  = content['mf']        
     
    #get MCs for the SU and MF only
    query =  {  '$and': [ { 'su_no': { '$eq': su } }, { 'mf_no' : { '$eq': mf } } ]  }          
    col = db["mcTable"]     
                    
    mc_array = []     
    results = col.find(query)
    for result in results:
        result["_id"] = str(result["_id"])
        mc_array.append(result) 
    
    session["mcTable"] = mc_array    

    if len(mc_array) > 0: 
        return  jsonify(mc_array), 200    
    

def establishSessionData():

    sessionData={}     

    email =""
   
    if (os.environ['ENVIRONMENT']=="PROD"):
        email = session['email']        
    else:       
        email = "vincent.cheng@macys.com" 
        email = "su123mf123@outlook.com" 
             
    col = db["userProfile"]

    query =  { "email": email}

    results = col.find_one(query)
    #print('results', results["email"])
    # print('results', results["mf_list"])
    
    #this forces the mf_list to be generated from profile only, not API requests.   
    
    session["userName"] = f"{results['first_name']} {results['last_name']}"
    session["mfList"] = results["mf_list"]   
        
    sessionData["userProfile"] = {"email" : results["email"], "first_name" : results["first_name"], "ignore_submit": results["ignore_submit"],
    "environment":  os.environ["ENVIRONMENT"], 
    "databaseSchema":  "dev" if database[:3].lower() == "dev" else "prod"    
     }    

    sessionData["mfList"] = results["mf_list"]   

    # get MC Table 
    # search = []        
    # for _filter in session['mfList']:
    #     search.append(   {  '$and': [ { 'su_no': { '$eq': _filter['SU'] } }, { 'mf_no' : { '$eq': _filter['MF'] } } ]  } )         
    
    # col = db["mcTable"]

    # query = {'$or' : search}        
    #                
    # mc_array = []     
    # results = col.find(query)
    # for result in results:
    #     result["_id"] = str(result["_id"])
    #     mc_array.append(result)
    #      
    ##sessionData["mcTable"] = mc_array
    ## session["mcTable"] = mc_array    

    sessionData["mcTable"] = ""  
    session["mcTable"] = ""  

    #get Party Table 
    search = []        
    for _filter in session['mfList']:
        search.append(   {  '$or': [ { '_id': { '$eq': _filter['SU'] } }, { '_id' : { '$eq': _filter['MF'] } } ]  } )         

    col = db["partyTable"]
    query = {'$or' : search}   
            
    party = []     
    results = col.find(query)
    for result in results:
        result["_id"] = str(result["_id"])
        party.append(result)

       
    partyTable = []     
    for pair in session['mfList']:
        for x in party:
            if x['_id'] == pair['SU']:
                pair['SU_NAME'] = x['party_name']
            if x['_id'] == pair['MF']:
                pair['MF_NAME'] = x['party_name']
        partyTable.append(pair)             

    sessionData["partyTable"] = partyTable     
    #print(partyTable)        


    col = db["qcAQL"]
    query = {}   
    aqlTable = []     
    results = col.find(query)
    for result in results:        
        aqlTable.append(result)
    
    sessionData["aqlTable"] = aqlTable


    col = db["checkList"]
    query = {}   
    checkList = []     
    results = col.find(query)
    for result in results:    
        #remove this _id as this is an object not serializable 
        result.pop('_id')    
        checkList.append(result)
    
    sessionData["checkListTemplate"] = checkList 
              
    #print("Established Session Data")


    # get SharePoint site - https://macysinc.sharepoint.com/sites/OSO/_api/Web/siteusers

    try: 
        _request = RequestOptions("{0}/_api/web/siteusers".format(site_url))
        _response = ctx.execute_request_direct(_request)
        _content_all = json.loads(_response.content)
        siteUsers = []
        for item in _content_all['d']['results']: 
            _username = item['Title']
            _email = item['Email']
            _id = item['Id']
            siteUser = { 
                    "id":_id,                
                    "username": _username,
                    "email": _email               
                } 
            siteUsers.append(siteUser)
    
    except Exception as e:        
        sessionData = None        
        print(f"***Error getting site user list from SharePoint***, {str(e)}")
        return sessionData

    session["siteUsers"] = siteUsers     
    sessionData["sharePointPath"] = os.environ['SHAREPOINT_PATH']
    session["sharePointReport"] = os.environ['SHAREPOINT_REPORT']

    #get QA members list     

    col = db["metaTable"]
    query = {'category': "qaList"}
    results = col.find_one(query)
 
    group = []
    for rec in results['selectionList']:          
        lead_email = rec["QALead"]
        for qa_email in rec['QAList']:            
            if qa_email["mqa"] == email:                
                group = rec['QAList']   

    sessionData["mqaMembers"] = group   

    #get Pack Type  
    col = db["metaTable"]
    query = {'category': "packType"}
    results = col.find_one(query)
    packTypes = results['selectionList']     
    sessionData["packTypes"] = packTypes     


    #get Product Category list     
    col = db["metaTable"]
    query = {'category': "productCategory"}
    results = col.find_one(query)
    productCategories = results['selectionList']     
    sessionData["productCategories"] = productCategories


    #get Inspection Type 
    col = db["metaTable"]
    query = {'category': "inspType"}
    results = col.find_one(query)    
    sessionData["inspType"] = results['selectionList']        
        
    return sessionData

####################################################################################
#  Genearte Excel Report - Start 
####################################################################################

# return distinct value of psGroupKey from the item psDictKey in a dictionary psDict
# psDict = Data Dictionary
# psDictKey = Dictionary item to get sum data
# psGroupKey = Distinct value 
def unique(psDict, psDictKey, psGroupKey):
    resultLst = []
    if psDictKey in psDict:
        try:
            keyfunc = lambda t: (t[psGroupKey])
            psDict[psDictKey].sort(key = keyfunc)
            for key, rows in itertools.groupby(psDict[psDictKey], keyfunc):
                resultLst.append (key)
        except:
            resultLst = []
    return resultLst

# return sum of psSumKey, group by psGroupKey, for the item psDictKey in a dictionary psDict.
# if no group by is needed, psGroupKey = ""
# psDict = Data Dictionary
# psDictKey = Dictionary item to get sum data
# psGroupKey = Item to be group by
# psSumKey = Item to sum 
def groupsum(psDict, psDictKey, psGroupKey, psSumKey):
    resultLst = []
    if psDictKey in psDict:
        if len(psGroupKey) > 0:
            try:
                keyfunc = lambda t: (t[psGroupKey])
                psDict[psDictKey].sort(key = keyfunc)

                for key, rows in itertools.groupby(psDict[psDictKey], keyfunc):
                    sumResult = (key, sum(r[psSumKey] 
                    for r in rows))
                    resultLst.append (sumResult)
            except:
                resultLst.append (('',0))
        else:
            try:
                sumResult = 0
                for i in psDict[psDictKey]:
                    sumResult += i[psSumKey]
                resultLst.append(sumResult)
            except:
                resultLst.append(0)
    else:
        if len(psGroupKey) > 0:
            resultLst.append ('',0)
        else:
            resultLst.append(0)
    return resultLst

# Generate report in Excel format
# version 18
# psWS = Excel worksheet
# psRptDict = Report Content Dictionary
# psRptFormat = Report Format Dictiionary
#   *if the item in psRptDict is dictionary, it will treated as a repeating items, and will populate across columns or rows based on the value stated in 
#    "nextRecord" in the psRptFormat
#   *if the item in psRptDict is not a dictionary, the value will be entered directly to the cell position.
#   *if header (left, center, right) and footer(left, center, right) has been specified in the psRptFormat, it will entered directly to the cell position
#    if not specific in the psRptFormat, it will use the Excel Header and Footer under page setup.
def genReport(psWS, psRptDict, psRptFormat):
    for lstKey, lstValue in psRptDict.items():
        try:
            if lstKey in psRptFormat["cell"]:
                if (isinstance(lstValue, dict)) == False:
                    psWS[(psRptFormat["cell"][lstKey])].value = lstValue
                else:
                    for lstKey2, lstValue2 in lstValue.items(): 
                        if len(lstValue2) > 0:
                            i = 0
                            j = 0
                            while i < len(lstValue2):
                                if lstKey2 in psRptFormat["cell"][lstKey]:
                                    if "nextRecord" in psRptFormat["cell"][lstKey]:
                                        if psRptFormat["cell"][lstKey]["nextRecord"] == "Column":
                                            mcell = True
                                            while mcell == True:
                                                colId = column_index_from_string(coordinate_from_string(psRptFormat["cell"][lstKey][lstKey2])[0]) + j
                                                rowId = coordinate_from_string(psRptFormat["cell"][lstKey][lstKey2])[1]
                                                if not isinstance(psWS.cell(row=rowId, column=colId), MergedCell):
                                                    mcell = False
                                                else:
                                                    mcell = True
                                                    j += 1                                                 
                                        else:
                                            mcell = True
                                            while mcell == True:
                                                colId = column_index_from_string(coordinate_from_string(psRptFormat["cell"][lstKey][lstKey2])[0])
                                                rowId = coordinate_from_string(psRptFormat["cell"][lstKey][lstKey2])[1] + i
                                                if not isinstance(psWS.cell(row=rowId, column=colId), MergedCell):
                                                    mcell = False
                                                else:
                                                    mcell = True
                                                    j += 1
                                        psWS.cell(row=rowId, column=colId, value=lstValue2[i])
                                i += 1
                                j += 1 
            else:
                if lstKey == "rightHeader":
                    psWS.HeaderFooter.oddHeader.right.text = lstValue
                    psWS.HeaderFooter.evenHeader.right.text = lstValue
                elif lstKey == "leftHeader":
                    psWS.HeaderFooter.oddHeader.left.text = lstValue
                    psWS.HeaderFooter.evenHeader.left.text = lstValue
                elif lstKey == "centerHeader":
                    psWS.HeaderFooter.oddHeader.center.text = lstValue
                    psWS.HeaderFooter.evenHeader.center.text = lstValue
                elif lstKey == "rightFooter":
                    psWS.HeaderFooter.oddFooter.right.text = lstValue
                    psWS.HeaderFooter.evenFooter.right.text = lstValue
                elif lstKey == "leftFooter":
                    psWS.HeaderFooter.oddFooter.left.text = lstValue
                    psWS.HeaderFooter.evenFooter.left.text = lstValue
                elif lstKey == "centerFooter":
                    psWS.HeaderFooter.oddFooter.center.text = lstValue
                    psWS.HeaderFooter.evenFooter.center.text = lstValue
                elif lstKey == "expandRow":
                    #expandRow format : [checking cell, start row, end row]
                    for item in lstValue:
                            if psWS[item[0]].value:
                                colId = column_index_from_string(coordinate_from_string(item[0])[0])
                                rowId = item[1]
                                while rowId <= item[2]:
                                    if psWS.cell(row=rowId, column=colId).value:
                                        psWS.row_dimensions[rowId].hidden = False
                                    rowId += 1
                elif lstKey == "expandCol":
                    #expandCol format : [checking cell, start columnm, end column]
                    for item in lstValue:
                            if psWS[item[0]].value:
                                rowId = coordinate_from_string(item[0])[1]
                                cellPos = item[1] + str(rowId)
                                colId = column_index_from_string(coordinate_from_string(cellPos)[0])
                                cellPos = item[2] + str(rowId)
                                colIdEnd = column_index_from_string(coordinate_from_string(cellPos)[0])
                                while colId <= colIdEnd:
                                    psWS.column_dimensions[get_column_letter(colId)].hidden = False
                                    colId += 1

                else:
                    1 == 1
        
        except:
            print ("Error found")
            return "Error", 701
    return "OK", 0    
 


#Version 30  7/18/22
@app.route('/printreport',methods=['POST'])
@check_logged
def printreport():     

     #fs = gridfs.GridFS(db)

    #_id of file example = "60b609a291e07bc60a594482"
    # below download_name and mimetype are dummy to fullfill the send_file requirements
    # changed from download_name to attachment_filename      
    # g=io.BytesIO(fs.get(ObjectId("6136da15d493601594076efc")).read())

    
    content = json.loads(request.headers['inspectionID'])
    inspectionID = content["inspectionID"]   
    
    localPrintTime = json.loads(request.headers['localTime'])
 

    inspMcno = inspectionID["mc"]
    inspIter = inspectionID["iteration"]
    inspType = inspectionID["type"]

    #print("Printing inspection report with inspection ID : {0}-{1}-{2} ".format(inspMcno, inspIter, inspType))

    

    # wb = load_workbook(filename=BytesIO(fs.get(ObjectId("6188b79dbc5219fc0944d8a9")).read()))
    # print('Generating inspeciton report...')
    # ws = wb["Inspection Report (Soft)"]
    # #can use either one of the below:
    # #ws.cell(row=2, column=3).value = 'Vnibcebt'
    # ws['C2'] = 'Yeah Yeah'               

    try:

        #db = cluster["qcDB"]
        colname = db["inspectionResult"]       
        colnameParty = db["partyTable"]        
        colnameExcelMap = db["fileDirectory"]
        colnameDefLst = db["defectTable"]
        # new added colnameUser below on 3/2/2022:
        colnameUser = db["userProfile"]
        # new added colnameUser below on 3/3/2022:
        colnameMeta = db["metaTable"]

        # inspMcno = "225461"
        # inspIter = "1"
        # inspType = "F"

        query =    { 'category' : { '$eq': "inspType" } }
        results = colnameMeta.find_one(query)
        typeArray  = results['selectionList']       
        inspTypeArray = filter(lambda x: x['insp_type'] == inspType, typeArray)
        inspTypeLong = list(inspTypeArray)[0]['insp_type_long']
       

        inspRecord = colname.find_one( { "_id.mc" : inspMcno,
                                "_id.iteration" : inspIter,
                                "_id.type" : inspType
                            }
        )

        defectList = list(colnameDefLst.find({ "lang" : 'en'}))

        # Determine MQA or SQA Report
        # If inspection record has been submitted, offical report will be use, else will use draft report
        # 3/2/2022 : Also to determine the Inspector Name based on email 
        if (inspRecord["misc"].get("qa_type", "").lower() == "mqa") and (inspRecord["misc"].get("submitted", False) == True):
            rpt = colnameExcelMap.find_one ( { "_id.excelFile": "InspRpt-MQA"} )
            InspBy = "MMGQA"
            inspEmail = inspRecord["misc"].get("mqa", False)
        elif (inspRecord["misc"].get("qa_type", "").lower() == "mqa") and (inspRecord["misc"].get("submitted", False) == False):
            rpt = colnameExcelMap.find_one ( { "_id.excelFile": "InspRpt-Draft-MQA"} )
            InspBy = "MMGQA"
            inspEmail = inspRecord["misc"].get("mqa", False)
        elif (inspRecord["misc"].get("qa_type", "").lower() == "sqa") and (inspRecord["misc"].get("submitted", False) == True):
            rpt = colnameExcelMap.find_one ( { "_id.excelFile": "InspRpt-SQA"} )
            InspBy = "SQA"
            inspEmail = inspRecord["misc"].get("sqa", False)
        else:
            rpt = colnameExcelMap.find_one ( { "_id.excelFile": "InspRpt-Draft-SQA"} )
            InspBy = "SQA"
            inspEmail = inspRecord["misc"].get("sqa", False)

        if inspEmail != False:
            userList = colnameUser.find_one( { "email" : inspEmail })
            if userList != None:
                inspName = userList["first_name"] + " " + userList["last_name"]
            else:
                inspName = ""
        else:
            inspName = ""


        #filename when using in Heroku:
        fs = gridfs.GridFS(db)
        wb = load_workbook(filename=BytesIO(fs.get(ObjectId(rpt["file"]["fileObj"])).read()))

        # filename in development:
        #wb = load_workbook(filename=rpt["file"]["fileName"])

        ws = wb[rpt["file"]["wsName"]]

        inspID = inspMcno + "-" + inspIter + "-" + inspType

        #Prepare figures for Order No, Order Qty and Ship Qty
        orderQtyLst = groupsum(inspRecord, "itemsTotal", "po_no", "order_qty")
        shipQtyLst = groupsum(inspRecord, "itemsTotal", "po_no", "ship_qty")
        qtyDict = {"poNo": [], "orderQty": [], "shipQty": []}
        qtyDict2 = {"poNo": [], "orderQty": [], "shipQty": []}
        i = 0
        for orderQtyItem in orderQtyLst:
            i += 1
            for shipQtyItem in shipQtyLst:
                if orderQtyItem[0] == shipQtyItem[0]:
                    if i <= 10:
                            qtyDict["poNo"].append(shipQtyItem[0])
                            qtyDict["orderQty"].append(orderQtyItem[1])
                            qtyDict["shipQty"].append(shipQtyItem[1])
                    else:
                            qtyDict2["poNo"].append(shipQtyItem[0])
                            qtyDict2["orderQty"].append(orderQtyItem[1])
                            qtyDict2["shipQty"].append(shipQtyItem[1])   


        # Prepare figures for Critical, Major, Minor, Total Defect, Accept Level, Reject Level and Visual Result
        criticalDefect = 0
        majorDefect = 0
        minorDefect = 0
        totalDefect = 0
        acceptLevel = 0
        rejectLevel = 0
        visualResult = ""
        if "visual_result" in inspRecord["misc"]:
            defectString = inspRecord["misc"]["visual_result"].get("defective_result", "")
            thresholdString = inspRecord["misc"]["visual_result"].get("sample_threshold", "")
            visualResult = inspRecord["misc"]["visual_result"].get("visual_result", "")
            if len(defectString) > 0:
                criticalDefect = int(defectString [:defectString.find("/")])
                majorDefect = int(defectString[defectString.find("/")+1 :defectString.find("/", defectString.find("/")+1)])
                minorDefect = int(defectString[-((len(defectString) -1 ) - defectString.find("/", defectString.find("/")+ 1)):])
                totalDefect = criticalDefect + majorDefect + math.floor(minorDefect / 4)

            if len(thresholdString) > 0:
                acceptLevel = int(thresholdString[:thresholdString.find("/")])
                rejectLevel = int(thresholdString[-((len(thresholdString) - 1) - thresholdString.find("/")):])                    

        # Prepare data for each checklist items.  chkDict1 in left section, ckhDict2 in middle and chkDict3 in the right section
        chkDict1 = { "result": []}
        chkDict2 = { "result": []}
        chkDict3 = { "result": []}
        chkDict4 = { "result": []}
        chkDict5 = { "result": []}

        k1 = rpt["cell"]["chkDict1"]["rptPosStart"]
        while k1 <= rpt["cell"]["chkDict1"]["rptPosEnd"]:
            item = next ((x for x in inspRecord["checkList"] if x["rpt_position"] == k1), "")
            if item != "":
                if item["result"].lower() != "select":
                    chkDict1["result"].append(item["result"])
                else:
                    chkDict1["result"].append(" ")
            else:
                chkDict1["result"].append(" ")
            k1 += 1
        k2 = rpt["cell"]["chkDict2"]["rptPosStart"]
        while k2 <= rpt["cell"]["chkDict2"]["rptPosEnd"]:
            item = next ((x for x in inspRecord["checkList"] if x["rpt_position"] == k2), "")
            if item != "":
                if item["result"].lower() != "select":
                    chkDict2["result"].append(item["result"])
                else:
                    chkDict2["result"].append(" ")
            else:
                chkDict2["result"].append(" ")
            k2 += 1
        k3 = rpt["cell"]["chkDict3"]["rptPosStart"]
        while k3 <= rpt["cell"]["chkDict3"]["rptPosEnd"]:
            item = next ((x for x in inspRecord["checkList"] if x["rpt_position"] == k3), "")
            if item != "":
                if item["result"].lower() != "select":
                    chkDict3["result"].append(item["result"])
                else:
                    chkDict3["result"].append(" ")
            else:
                chkDict3["result"].append(" ")
            k3 += 1
        k4 = rpt["cell"]["chkDict4"]["rptPosStart"]
        while k4 <= rpt["cell"]["chkDict4"]["rptPosEnd"]:
            item = next ((x for x in inspRecord["checkList"] if x["rpt_position"] == k4), "")
            if item != "":
                if item["result"].lower() != "select":
                    chkDict4["result"].append(item["result"])
                else:
                    chkDict4["result"].append(" ")
            else:
                chkDict4["result"].append(" ")
            k4 += 1
        k5 = rpt["cell"]["chkDict5"]["rptPosStart"]
        while k5 <= rpt["cell"]["chkDict5"]["rptPosEnd"]:
            item = next ((x for x in inspRecord["checkList"] if x["rpt_position"] == k5), "")
            if item != "":
                if item["result"].lower() != "select":
                    chkDict5["result"].append(item["result"])
                else:
                    chkDict5["result"].append(" ")
            else:
                chkDict5["result"].append(" ")
            k5 += 1

        #Get Defect data - use product, defect code and defect detail code to get the english defect description
        defectDict = { "defect":[], "critical":[], "major":[], "minor": []}
        for defectItem in inspRecord["defects"]:
            dItem = [p for p in defectList if p["product"] == defectItem.get("product") and p["defect_code"] == defectItem.get("defect_area_code") and p["defect_detail_code"] == defectItem.get("defect_detail_area")]
            if len(dItem) > 0:
                defectDict["defect"].append(dItem[0]["defect_code"] + "-" + dItem[0]["defect_area"] + " " + dItem[0]["defect_detail_code"] + "-" + dItem[0]["defect_detail"])
            else:
                defectDict["defect"].append(defectItem.get("defect_area_code", "") + "-" + defectItem.get("defect_detail_area"))
            
            if defectItem.get("critical", False) == True:
                defectDict["critical"].append(defectItem.get("major", ""))
                defectDict["major"].append(0)
            else:
                defectDict["critical"].append(0)
                defectDict["major"].append(defectItem.get("major", ""))
            defectDict["minor"].append(defectItem.get("minor", ""))

        # Get the Inspection Date
        if "inspection_date" in inspRecord["main"]:
            inspDate = inspRecord["main"]["inspection_date"]
        else:
            inspDate = "-"

        #Set Footname as Inspection ID + Print Date
        #footer = "Inspection ID: " + inspID + " \n " + "Print Date: " + datetime.now().strftime('%m/%d/%YYYY %H:%M:%S')
        footer = "Inspection ID: " + inspID + " \n " + "Print Date: " + localPrintTime
       
     
       # dictionary for the output.  key in this dictionary must match the key the collection: excelMapping
        report = {
        "suNo" : inspRecord["main"].get("su_no", " "),
        "mfNo" : inspRecord["main"].get("mf_no", " "),
        "suName": colnameParty.find_one ( { "_id" : inspRecord["main"].get("su_no", " ") } )["party_name"], 
        "mfName" : colnameParty.find_one ( { "_id" : inspRecord["main"].get("mf_no"," ") } )["party_name"], 
        "inspRecordNo": inspID,
        "brand" : inspRecord["misc"].get("label", " "),
        "inspType" : inspTypeLong,
        "inspDate": inspDate,
        "mcNo" : (",".join(unique(inspRecord, "itemsTotal", "mc_no"))),
        "style" : (",".join(unique(inspRecord, "itemsTotal", "item_no"))),
        "prodType" : inspRecord["misc"].get("product_type", " "),
        "packType" : inspRecord["misc"].get("pack_type", " "),
        "fibreContent" : inspRecord["misc"].get("fibre_content", " " ),
        "shipMode" : inspRecord["misc"].get("ship_mode", " "),
        "labTestReport" : inspRecord["misc"].get("lab_test_report", " "),
        "childSafetyReport" : inspRecord["misc"].get("lab_child_safety_report", " "),
        "prop65Report" : inspRecord["misc"].get("props65_report", " "),
        "packPct" : inspRecord["misc"].get("pack_pct", " "),
        "shipWindow" : inspRecord["misc"].get("ship_window"," "),
        "inspector" : inspName,
        "qtyDict": qtyDict,
        "qtyDict2" : qtyDict2,
        "inspQty": groupsum(inspRecord, "itemsTotal", "", "inspect_qty")[0],
        "acceptLevel": acceptLevel,
        "rejectLevel": rejectLevel,
        "criticalDefect": criticalDefect,
        "majorDefect": majorDefect,
        "minorDefect": minorDefect,
        "totalDefect": totalDefect,
        "visualResult": visualResult,
        "packingResult": inspRecord["misc"].get("packing_result", ""),
        "measureResult": inspRecord["misc"].get("measurement_result", ""),
        "finalResult": inspRecord["misc"].get("final_result", ""),
        "chkDict1": chkDict1,
        "chkDict2": chkDict2,
        "chkDict3": chkDict3,
        "chkDict4": chkDict4,
        "chkDict5": chkDict5,
        "defectDict": defectDict,
        "comments": inspRecord["misc"].get("comments", ""),
        "cartonList": inspRecord["misc"].get("carton_list", ""),
        "cartonTotal": inspRecord["misc"].get("carton_total", ""),
        "rightHeader": InspBy,
        "leftFooter": footer,
        "expandRow" : [["C24", 23, 26], ["B57", 57, 68]],
        "expandCol": [["T2", "T", "V"], ["X3", "Y", "Z"]]

        }



        result = genReport(ws, report, rpt)

        #Output in Heroku:        

        ws.protection.selectLockedCells=True
        ws.protection.selectUnlockedCells=True
        ws.protection.objects=True
        ws.protection.scenarios=True 
        
        ws.protection.sheet = True
        ws.protection.enable()
        ws.protection.password="mismis"


        out = BytesIO()
        wb.save(out)
        out.seek(0)

        #Output in development:
        #wb.save(filename="C:\Temp\MongoDB\InspRpt2.xlsx")

        wb.close()
        ## if no error, result = ("OK", 0), if there is error result = ("Error", 701)


    except Exception as e:
            print("Failed to convert into Excel format")
            print(str(e))

    finally:   
            print("Printing Inspection Result - Print Complete")                                      
            #wb.close()    
        
    return send_file(out,  attachment_filename='inspection_report.xls', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


####################################################################################
#  Genearte Excel Report - End
####################################################################################

###################################################################################
#  SharePoint Section - Start
####################################################################################


@app.route('/api/getsharepointfiles',methods=['POST'])
@check_logged
def sharepointfiles():       

    try:        
        content = request.get_json() #python data     
        folder = content['folder']
        inspection_id = content['inspection_id']
        
        # ## below are for testing 
        sharePointReport = "2022 Inspection Report"
        relative_url = "2022InspRpt" + "/" + folder

        ## below are getting from environment 
        # sharePointReport = os.environ['SHAREPOINT_REPORT']
        # relative_url = os.environ['SHAREPOINT_PATH'] + "/" + folder        
                
        libraryRoot = ctx.web.get_folder_by_server_relative_path(relative_url)

        ctx.load(libraryRoot)
        ctx.execute_query()

        #if you want to get the files in the folder        
        files = libraryRoot.files
        ctx.load(files)
        ctx.execute_query()

        for file in files:    
            _name = file.properties["Name"]    
            #print("Folder {0}, File name: {1}".format(folder, _name))

        #if you want to get the items in the folder        
        caml_query = CamlQuery()
        caml_query.ViewXml = '''<View Scope="RecursiveAll"><Query><Where><Eq><FieldRef Name='Inspection_x0020_ID' /><Value Type='Text'>{0}</Value></Eq></Where></Query></View>'''.format(inspection_id)
        caml_query.FolderServerRelativeUrl = relative_url
    
        # 3 Retrieve list items based on the CAML query 
        #oList = ctx.web.lists.get_by_title('2022inspRpt') - title must match the list name in SharePoint
        oList = ctx.web.lists.get_by_title(sharePointReport) 
        items = oList.get_items(caml_query) 
        ctx.execute_query()

        sharePoint_array = []         
        for item in items:                
            _inspection_id = item.properties["Inspection_x0020_ID"]                    
            _modified = item.properties["Modified"]
            _author = item.properties["AuthorId"]
            _modified_by = item.properties["EditorId"]    
            _last_editor = item.properties["Last_x0020_Editor"]                    
            _id  = item.properties["Id"]                
            list_item  = item.expand(["File"])
            list_item = ctx.web.lists.get_by_title(sharePointReport).get_item_by_id(_id).expand(["File"])
            ctx.load(list_item)
            ctx.execute_query()           

            _editor = ""
            for user in session["siteUsers"]:
                if user['id'] == _modified_by:
                    _editor = user['username']         

            _size = list_item.file.properties['Length']                         
            _size = convert_size(int(_size))            

            sharePoint_items = { 
                "folder":folder,
                "filename": list_item.file.properties['Name'],
                "inspection_id": _inspection_id,
                "modified":_modified,
                "editor":_editor,
                "size": _size,
                "url": "https://macysinc.sharepoint.com" + list_item.file.properties["ServerRelativeUrl"],
                 "relative_path" : list_item.file.properties["ServerRelativeUrl"],
                 "unique_id": list_item.file.unique_id,
                 "last_editor": _last_editor
            }             
            #print('id', list_item.file.unique_id)
            sharePoint_array.append(sharePoint_items)

        # if (sharePoint_array == []): 
        #     ## No files on SharePoint
        #     sharePoint_items = {
        #         "folder":None, 
        #         "filename": 'No files',
        #         "inspection_id": None,
        #         "modified":datetime.now(timezone.utc),
        #         "editor":None,
        #         "size": None,
        #         "url" :None,
        #         "relative_path" : None,
        #         "unique_id": None,
        #         "last_editor": None 
        #     } 
        #     sharePoint_array.append(sharePoint_items)
       
        return  jsonify(sharePoint_array), 200         

    except:
        print('error')
        #return "Error", 404    

#######  SharePoint Combine photos to PDF
##  /api/createsharePointPDF

@app.route('/api/createsharepointpdf',methods=['POST'])
@check_logged
def createsharepointpdf():       

    try:              
        content = request.get_json() #python data     
        folder = content['folder']
        inspection_id = content['inspection_id']
        # inspection_id = request.headers['inspection_id']
        # relative_url =  request.headers['relative_url']
        inspection_date =  request.headers['inspection_date']
       
        # below are okay 
        # sharePointReport = "9999 Inspection Report"
        # relative_url = "9999InspRpt" + "/" + folder

        ## below are getting from environment 
        sharePointReport = os.environ['SHAREPOINT_REPORT']
        relative_url = os.environ['SHAREPOINT_PATH'] + "/" + folder        
                
        libraryRoot = ctx.web.get_folder_by_server_relative_path(relative_url)

        ctx.load(libraryRoot)
        ctx.execute_query()

        #if you want to get the files in the folder        
        files = libraryRoot.files
        ctx.load(files)
        ctx.execute_query()

        # for file in files:    
        #     _name = file.properties["Name"]    
            #print("Folder {0}, File name: {1}".format(folder, _name))

        #if you want to get the items in the folder        
        caml_query = CamlQuery()
        caml_query.ViewXml = '''<View Scope="RecursiveAll"><Query><Where><Eq><FieldRef Name='Inspection_x0020_ID' /><Value Type='Text'>{0}</Value></Eq></Where></Query></View>'''.format(inspection_id)
        caml_query.FolderServerRelativeUrl = relative_url
    
        # 3 Retrieve list items based on the CAML query 
        #oList = ctx.web.lists.get_by_title('2022inspRpt') - title must match the list name in SharePoint
        oList = ctx.web.lists.get_by_title(sharePointReport) 
        items = oList.get_items(caml_query) 
        ctx.execute_query()

        sharePoint_array = []         
        for item in items:                    
            _id  = item.properties["Id"]                
            list_item  = item.expand(["File"])
            list_item = ctx.web.lists.get_by_title(sharePointReport).get_item_by_id(_id).expand(["File"])
            ctx.load(list_item)
            ctx.execute_query()                    

            if allowed_photos(list_item.file.properties['Name']):
                sharePoint_items = { 
                    "filename": list_item.file.properties['Name'],
                    "url": "https://macysinc.sharepoint.com" + list_item.file.properties["ServerRelativeUrl"],
                    "relative_path" : list_item.file.properties["ServerRelativeUrl"],
                    "unique_id": list_item.file.unique_id
                }             
                #print('id', list_item.file.unique_id)
                sharePoint_array.append(sharePoint_items)

        if (sharePoint_array == []): 
            return "No photos for PDF generation !", 405   

        #-------------------------------------------------------
        # Preparing zip file
        #-------------------------------------------------------
        
        
        archive = BytesIO()
        with ZipFile(archive, 'w') as zip_archive:

            for rec in sharePoint_array:
                     
                file_url =  rec['relative_path']
                _response = File.open_binary(ctx, file_url)
                data = BytesIO(_response.content) 
                
                # convert to dpi=300 and 400 pixels = 400px /72 = 5.5 inches on PDF A4 portait 
                basewidth = 600 
                img = Image.open(data) 
                wpercent = (basewidth/float(img.size[0]))
                hsize = int((float(img.size[1])*float(wpercent)))
                img = img.resize((basewidth,hsize), Image.ANTIALIAS)
                data2 = BytesIO()    
                img.save(data2, dpi=(300,300),format=image_format(rec['filename']))
                ##nim = Image.open(data2)     
                
                with zip_archive.open(rec['filename'], 'w') as _file:                                        
                    print(f"processing {rec['filename']} with extension {rec['filename'].rsplit('.', 1)[1].upper()}")
                    imageFormat = image_format(rec['filename'])                             
                    _file.write(data2.getvalue())        
                _file.close

        #-------------------------------------------------------
        # Dumping PDF stream to file
        #-------------------------------------------------------
       
        out = BytesIO()            
        out.write(archive.getbuffer())                       
        out.seek(0)    
        archive.close()

        #print("passing PDF")
        target_folder = ctx.web.get_folder_by_server_relative_path(relative_url)
        ctx.execute_query()

        info = FileCreationInformation()

        info.content = out.read()

        # Give the name of the photo albumn  
        filename = inspection_id + " Photo Album.zip" 

        info.url = filename  
        
        info.overwrite = True
        upload_file = target_folder.files.add(info)
        ctx.execute_query()

        #Once file is uploaded, it's metadata could be set like this
            
        list_item = upload_file.listItemAllFields # get associated list item 
        list_item.set_property("Inspection_x0020_ID", inspection_id)
        list_item.set_property("Inspection_x0020_Date", inspection_date)
        list_item.set_property("Last_x0020_Editor", session["userName"])
 
        list_item.update()
        ctx.execute_query()
        
        #-------------------------------------------------------
        # Ending zip file
        #-------------------------------------------------------
        
        #-------------------------------------------------------
        # preparing a new PDF 
        #-------------------------------------------------------
        
        pdf = Document()
        #ttf = ImageFont.truetype("courbi.ttf", 20)    
                                    
        for rec in sharePoint_array:

            #print('SharePoint files : ', rec['relative_path'])     
            
            page = Page()
            pdf.append_page(page)
            page_layout = SingleColumnLayout(page)

            
            file_url =  rec['relative_path']
            _response = File.open_binary(ctx, file_url)
            data = BytesIO(_response.content) 

            # convert to dpi=300 and 400 pixels = 400px /72 = 5.5 inches on PDF A4 portait 
            basewidth = 600 
            img = Image.open(data) 
            wpercent = (basewidth/float(img.size[0]))
            hsize = int((float(img.size[1])*float(wpercent)))
            img = img.resize((basewidth,hsize), Image.ANTIALIAS)
            data2 = BytesIO()    
            img.save(data2, dpi=(300,300),format=image_format(rec['filename']))
            nim = Image.open(data2)                                            
                                                
            page_layout.add(Paragraph(rec['filename'], horizontal_alignment=Alignment.CENTERED))
            page_layout.add(
                BorbImage(nim, width=Decimal(400),  height=Decimal(400),  horizontal_alignment=Alignment.CENTERED )
            )              

        #-------------------------------------------------------
        # Dumping PDF stream to file
        #-------------------------------------------------------

        #print("passing PDF")
        out = BytesIO()							
        PDF.dumps(out, pdf)
        out.seek(0)

        #inspection_id = '225233-1-F' 
        #relative_url =  "2022inspRpt/SU22975MF36843"
        
        target_folder = ctx.web.get_folder_by_server_relative_path(relative_url)
        ctx.execute_query()

        info = FileCreationInformation()

        info.content = out.read()

        # Give the name of the photo albumn  
        filename = inspection_id + " Photo Album.pdf" 

        info.url = filename  
        
        info.overwrite = True
        upload_file = target_folder.files.add(info)
        ctx.execute_query()

        #Once file is uploaded, it's metadata could be set like this
            
        list_item = upload_file.listItemAllFields # get associated list item 
        list_item.set_property("Inspection_x0020_ID", inspection_id)
        list_item.set_property("Inspection_x0020_Date", inspection_date)
        list_item.set_property("Last_x0020_Editor", session["userName"])
 
        list_item.update()
        ctx.execute_query()

        #Once all including metadata update is done, proceed to delete the photos.  
        for rec in sharePoint_array:
            #print('SharePoint files : ', rec['relative_path'])     
            f = ctx.web.get_file_by_id(rec['unique_id'])
            f.delete_object()
        ctx.execute_query()
        
        # return  jsonify(sharePoint_array), 200         
        return  "", 200         

    except Exception as e:
        print('error', e)
        return "Something went wrong !", 404    

#######  SharePoint delte photos from SharePoint
@app.route('/api/deleteSPfile',methods=['POST'])
@check_logged
def delete_sp_file():
        
    content = request.get_json() #python data     
    _id = content['_id'] # file id
     
    try:  

        f = ctx.web.get_file_by_id(_id)
        ##ctx.execute_query()
        f.delete_object()
        ctx.execute_query()
        
        return  "OK", 200 
      
    
    #except mongoengine.errors.OperationError:           
    except OperationFailure:
        print("error")
        return "Error", 400 



###################################################################################
#  SharePoint Section - End
####################################################################################


############################################################### 
#  Past X days inspection - Start
############################################################### 

@app.route('/api/pastXdaysResult',methods=['POST'])
@check_logged
def pastXdaysResult():
        
    content = request.get_json() #python data     
    #mc = content['mc']
    #print("mc", mc) 

    user =""
    if session.get('user'):   ##  production environment 
        user = session['user']['email']
        user = user.lower()               
    else:        
        user = "vincent.cheng@macys.com"

    ## build a mf_list for the current user. 
    col = db["userProfile"]
    query = {'email' : user}
    result = col.find(query) 
    mf_list = []
    for rec in result[0]['mf_list']:
        mf_list.append({'MF' : rec['MF']})

    col = db["inspectionResult"]
    search = []

    # Getting the 1st day of last month
    today = datetime2.date.today()
    first = today.replace(day=1)
    lastDayLastMonth = first - datetime2.timedelta(days=1)
    lastDay2MonthsAgo = lastDayLastMonth.replace(day=1)  - datetime2.timedelta(days=1)
    #firstLastMonth = lastMonth.replace(day=1)

    # print(firstLastMonth.strftime("%D"))
    # print(firstLastMonth.strftime("%Y%m"))

    #firstLastMonthISO = firstLastMonth.strftime("%Y-%m-%dT%H:%M:%SZ")     
    #firstLastMonthISOString = dateutil.parser.parse(firstLastMonthISO)

    lastDay2MonthsAgoISO =  lastDay2MonthsAgo.strftime("%Y-%m-%d")     

    for filter in mf_list:
        search.append(   {  '$and': [         
            { 'main.mf_no' : { '$eq': filter['MF'] } },
            { 'main.inspection_date' : { '$gt' : lastDay2MonthsAgoISO } }       
        ]  } )        
    
    query = {'$or' : search}
    #results = col.find(query).limit(15)
    results = col.find(query)

    query_sort = [  ("main.inspection_date",pymongo.DESCENDING), ("_id",pymongo.ASCENDING)]    
    #results = col.find( query_filter).sort(query_sort).limit(100)
    results = col.find(query).sort(query_sort)
    
    # query_filter =""
    # if session.get('user'):   ##  prod
    #     user = session['user']['email']
    #     user = user.lower()       
    #     query_filter =  {"misc.sqa" : user} if  "@macys.com" not in session.get('email').lower() else {'$and' : [  { 'misc.mqa' : { '$eq' : user }},  {'misc.qa_type' : { '$eq' : "MQA"}}]}        
    # else:        
    #     user = "henry.khut@macys.com"
    #     query_filter =  {"misc.mqa" : user} 

    inspection_array = []
    for result in results:    
         inspection_array.append(result)        
            
    if (results):
       return  jsonify({"pastInspections" : inspection_array, "user" : user}), 200 
    else:
       return "Error", 404 


############################################################### 
#  Past X days inspection - End
############################################################### 


def convert_size(size_bytes):
   if size_bytes == 0:
       return "0B"
   size_name = ("B", "KB", "MB", "GB", "TB", "PB", "EB", "ZB", "YB")
   i = int(math.floor(math.log(size_bytes, 1024)))
   p = math.pow(1024, i)
   s = round(size_bytes / p, 2)
   return "%s %s" % (s, size_name[i])


@app.route("/map-marker")
def map_marker():
    # this map using stamen terrain
    # we add some marker here
    print("loading the map")          
    #Position to the center of the map     
    f = folium.Figure(width="100%", height="100%")
    map = folium.Map(
        location=[29.81925 , 31.35978],              
        tiles='Stamen Terrain',
        ##zoom_start=2.8
        ##zoom_start=2.45
        zoom_start=2.3, min_zoom = 2.5,
        max_bounds=True
    )   
 
    # countries = [
    #     {"country":'Nicaragua', "lat": 12.769013, "long":-85.602364, "factories": 3},
    #     {"country":'Vietname', "lat": 21.028280, "long":105.853882, "factories": 5},
    #     {"country":'Italy', "lat": 43.769562, "long":11.255814, "factories": 2},
    #     {"country":'Shenzhen', "lat": 22.542883, "long":114.062996, "factories": 10}
    # ]
    countries = [
        {"country":'BANGLADESH', "lat": 23.4044, "long":90.3126, "factories": 5},
        {"country":'CAMBODIA', "lat": 11.358899, "long":104.73928, "factories": 8},
        {"country":'CHINA', "lat": 31.1382877, "long":121.34404, "factories": 154},
        {"country":'EGYPT', "lat": 29.81925, "long":31.35978, "factories": 5},
        {"country":'INDIA', "lat": 22.921431, "long":76.9802, "factories": 26},
        {"country":'INDONESIA', "lat": -6.442596, "long":107.448334, "factories": 16},
        {"country":'ITALY', "lat": 43.83173, "long":11.17973, "factories": 1},
        {"country":'JORDAN', "lat": 32.45555, "long":35.96866, "factories": 2},
        {"country":'KOREA', "lat": 37.2558, "long":127.100226, "factories": 2},
        {"country":'LESOTHO', "lat": -29.466667, "long": 27.933333, "factories":  2},
        {"country":'MADAGASCAR', "lat": -18.916667 , "long": 47.516667, "factories": 1},
        {"country":'MALAYSIA', "lat": 1.66944, "long":103.5787, "factories": 1},
        {"country":'NICARAGUA', "lat": 11.86909, "long":-86.10617, "factories": 2},
        {"country":'PAKISTAN', "lat": 24.8406018, "long":67.2272526, "factories": 9},
        {"country":'PHILIPPINES', "lat": 13.92763, "long":121.09305, "factories": 6},
        {"country":'POLAND', "lat": 49.70333, "long":21.73177, "factories": 1},
        {"country":'PORTUGAL', "lat": 41.41959, "long":-8.39006, "factories": 2},
        {"country":'SRI LANKA', "lat": 7.2981, "long":79.8976, "factories": 3},
        {"country":'TAIWAN', "lat": 25.066667, "long":121.51667, "factories": 1},
        {"country":'THAILAND', "lat": 13.524, "long":100.2905, "factories": 3},
        {"country":'TURKEY', "lat": 37.80651, "long":29.24568, "factories": 3},
        {"country":'VIETNAM', "lat": 19.033419, "long":105.580931, "factories": 30}
    ]


    for rec in countries:
        country = rec['country']
        lat = rec['lat']
        long = rec['long']
        factories = rec['factories']

        folium.Marker(
            location=[lat, long],                  
            popup="<b style='color:red'>" + country + ":<br>" + str(factories) + " active factories</b>",
            tooltip=country,
            icon=folium.Icon(color="red",background_color="white", icon="fa-star", prefix='fa')
        ).add_to(map)
       
    #return map._repr_html_()
    return render_template('show_map.html', map=map._repr_html_())
   

if __name__ == "__main__":

    if (os.environ['ENVIRONMENT']=="DEV"):
        app.static_folder = "frontend/public/static"
        app.template_folder = "frontend/public"     
    
    app.run(host='127.0.0.1', port=5000, debug=True)
